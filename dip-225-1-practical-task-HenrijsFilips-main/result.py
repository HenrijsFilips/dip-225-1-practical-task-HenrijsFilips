from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
total=0
max_row = ws.max_row
for row in range(2, max_row + 1):
    vards = ws['A' + str(row)].value
    stunda = ws['B' + str(row)].value
    likme = ws['C' + str(row)].value
    if isinstance(stunda, (int, float)) and isinstance(likme, (int, float)):
        if 'a' not in str(stunda):
            salary = stunda * likme
            ws['D' + str(row)] = salary
            print(round(salary, 2))

wb.save('tests/test1.xlsx')
wb.close()
